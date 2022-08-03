import openpyxl

# criando uma planilha (book)

book = openpyxl.Workbook()

# visualizando pgs existentes

print(book.sheetnames)

# criando um pg

book.create_sheet('Frutas')

# selecionando uma pg

frutas_page = book['Frutas']
frutas_page.append(['Fruta', 'Quantidade', 'Preço'])
frutas_page.append(['Banana', '5', 'R$3,90'])
frutas_page.append(['Pera', '2', 'R$5'])
frutas_page.append(['Goiaba', '5', 'R$1,20'])
frutas_page.append(['Manga', '5', 'R$2,40'])

# salvando a planilha

book.save('Planilha de preços frutas.xlsx')
